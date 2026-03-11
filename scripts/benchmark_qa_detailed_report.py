#!/usr/bin/env python3
import argparse
import json
import math
import os
import re
import time
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib import error, parse, request

import openpyxl


NO_ANSWER_PATTERNS: List[Tuple[re.Pattern[str], str]] = [
    (re.compile(r"i could not find relevant information in the available company documents\.?", re.I), "no_relevant_documents"),
    (re.compile(r"i can[’']?t confirm from the provided documents", re.I), "cannot_confirm_with_documents"),
    (re.compile(r"provided documents do not contain enough information", re.I), "insufficient_document_evidence"),
    (re.compile(r"answer generation failed due to a temporary model issue", re.I), "generation_failure"),
    (re.compile(r"利用可能な社内文書内で、要求された情報は見つかりませんでした"), "no_relevant_documents_ja"),
    (re.compile(r"提供された文書から確認できません"), "cannot_confirm_with_documents_ja"),
    (re.compile(r"回答生成に一時的な問題が発生しました"), "generation_failure_ja"),
]

NO_ANSWER_EXPLANATIONS: Dict[str, str] = {
    "no_relevant_documents": "Retriever did not find supporting documents for the question in indexed corpora.",
    "cannot_confirm_with_documents": "Model detected insufficient evidence in retrieved documents and refused to assert a fact.",
    "insufficient_document_evidence": "Retrieved context lacked enough detail to answer confidently.",
    "generation_failure": "LLM generation failed transiently during answer synthesis.",
    "no_relevant_documents_ja": "Retriever did not find relevant documents (Japanese response path).",
    "cannot_confirm_with_documents_ja": "Model could not confirm answer from retrieved documents (Japanese response path).",
    "generation_failure_ja": "LLM generation failed transiently (Japanese response path).",
    "empty_answer": "Task finished but answer body was empty.",
    "missing_result": "No result was recorded for this question.",
}


def has_cjk(text: str) -> bool:
    return bool(re.search(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]", text or ""))


def normalize_whitespace(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def normalize_header(s: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def get_cell_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return str(v).strip()
    if hasattr(v, "text"):
        return str(getattr(v, "text") or "").strip()
    return str(v).strip()


def tokenize(text: str) -> List[str]:
    t = normalize_whitespace(text).lower()
    if not t:
        return []
    if has_cjk(t):
        compact = re.sub(r"[^\w\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]+", "", t)
        if len(compact) < 2:
            return list(compact)
        return [compact[i : i + 2] for i in range(len(compact) - 1)]
    return [w for w in re.sub(r"[^a-z0-9\s]", " ", t).split() if len(w) > 1]


def jaccard(a: Iterable[str], b: Iterable[str]) -> float:
    sa = set(a)
    sb = set(b)
    if not sa and not sb:
        return 1.0
    inter = len(sa.intersection(sb))
    union = len(sa.union(sb))
    return float(inter / union) if union else 0.0


def overlap_recall(a: Iterable[str], b: Iterable[str]) -> float:
    sa = set(a)
    sb = set(b)
    if not sa:
        return 0.0
    inter = len(sa.intersection(sb))
    return float(inter / len(sa))


def now_ts() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def fmt_ms(v: Any) -> str:
    try:
        n = float(v)
    except Exception:
        return "N/A"
    if not math.isfinite(n) or n <= 0:
        return "N/A"
    return f"{int(round(n))} ms"


def fmt_sec(v: Any) -> str:
    try:
        n = float(v)
    except Exception:
        return "N/A"
    if not math.isfinite(n) or n <= 0:
        return "N/A"
    return f"{n / 1000.0:.2f} s"


def classify_no_answer(status: str, answer: str) -> str:
    if status != "FINISHED":
        return f"status_{status.lower()}"
    if not answer.strip():
        return "empty_answer"
    for rx, reason in NO_ANSWER_PATTERNS:
        if rx.search(answer):
            return reason
    return ""


def explain_reason(reason: str) -> str:
    if reason in NO_ANSWER_EXPLANATIONS:
        return NO_ANSWER_EXPLANATIONS[reason]
    if reason.startswith("status_timeout"):
        return "Task timed out before completion, likely queue delay or backend processing delay."
    if reason.startswith("status_failed"):
        return "Task ended in FAILED status; inspect API/worker logs for generation or retrieval errors."
    if reason.startswith("status_cancel"):
        return "Task was canceled before producing a final answer."
    if reason.startswith("task_create_error"):
        return "Could not enqueue task; request failed before retrieval/generation."
    if reason.startswith("kpi_fetch_error"):
        return "KPI lookup failed; analytics event may be delayed or missing."
    if reason.startswith("trace_fetch_error"):
        return "Trace lookup failed; performance tracing may be disabled or not sampled."
    if reason.startswith("eval_low_semantic"):
        return "Output semantic alignment against expected answer was low."
    if reason == "expected_source_not_cited":
        return "Answer did not cite the expected source from benchmark sheet."
    if reason == "rag_not_used":
        return "Pipeline returned with ragUsed=false, so retrieval-grounded response path was not used."
    if reason == "missing_source_citation":
        return "Answer had no explicit SOURCE/SOURCES block to ground the response."
    return "Unclassified issue; inspect row-level details."


def extract_answer(content: str) -> str:
    raw = str(content or "")
    m = re.search(r"<!--SINGLE_LANG_START-->([\s\S]*?)<!--SINGLE_LANG_END-->", raw)
    if m:
        try:
            obj = json.loads(m.group(1))
            return normalize_whitespace(obj.get("content") or "")
        except Exception:
            return normalize_whitespace(raw)
    m2 = re.search(r"<!--DUAL_LANG_START-->([\s\S]*?)<!--DUAL_LANG_END-->", raw)
    if m2:
        try:
            obj = json.loads(m2.group(1))
            return normalize_whitespace(obj.get("translated") or obj.get("japanese") or "")
        except Exception:
            return normalize_whitespace(raw)
    return normalize_whitespace(raw)


def extract_sources(answer: str) -> List[str]:
    lines = [ln.strip() for ln in str(answer or "").splitlines()]
    out: List[str] = []
    for ln in lines:
        if re.match(r"^SOURCE\s*:", ln, re.I):
            out.append(re.sub(r"^SOURCE\s*:\s*", "", ln, flags=re.I).strip())
    for i, ln in enumerate(lines):
        if re.match(r"^SOURCES?\s*:", ln, re.I):
            for j in range(i + 1, len(lines)):
                x = lines[j].strip()
                if not x:
                    continue
                if x.startswith("- "):
                    out.append(x[2:].strip())
                elif re.match(r"^[A-Za-z]+\s*:", x):
                    break
    deduped: List[str] = []
    seen = set()
    for x in out:
        if x and x not in seen:
            seen.add(x)
            deduped.append(x)
    return deduped


def has_inline_citations(answer: str) -> bool:
    return bool(re.search(r"\[\d{1,2}\]", answer or ""))


def evaluate_answer(expected: str, actual: str, expected_source: str, actual_sources: List[str]) -> Dict[str, Any]:
    expected_clean = normalize_whitespace(expected)
    actual_clean = normalize_whitespace(actual)
    source_clean = normalize_whitespace(expected_source)
    if not expected_clean:
        source_match = "N/A"
        if source_clean:
            hay = " ".join(actual_sources + [actual_clean]).lower()
            source_match = "YES" if source_clean.lower() in hay else "NO"
        return {
            "recall_pct": None,
            "jaccard_pct": None,
            "semantic_pct": None,
            "verdict": "NO_EXPECTED_ANSWER",
            "source_match": source_match,
        }

    exp_tokens = tokenize(expected_clean)
    out_tokens = tokenize(actual_clean)
    recall_pct = int(round(overlap_recall(exp_tokens, out_tokens) * 100))
    jaccard_pct = int(round(jaccard(exp_tokens, out_tokens) * 100))
    semantic_pct = int(round((recall_pct * 0.6) + (jaccard_pct * 0.4)))

    if semantic_pct >= 75:
        verdict = "PASS"
    elif semantic_pct >= 45:
        verdict = "PARTIAL"
    else:
        verdict = "FAIL"

    source_match = "N/A"
    if source_clean:
        hay = " ".join(actual_sources + [actual_clean]).lower()
        source_match = "YES" if source_clean.lower() in hay else "NO"

    return {
        "recall_pct": recall_pct,
        "jaccard_pct": jaccard_pct,
        "semantic_pct": semantic_pct,
        "verdict": verdict,
        "source_match": source_match,
    }


def summarize_stage_list(trace: Optional[Dict[str, Any]]) -> List[str]:
    if not trace or not isinstance(trace, dict):
        return []
    stages = trace.get("stages")
    if not isinstance(stages, list):
        return []
    out = []
    for s in stages:
        if not isinstance(s, dict):
            continue
        name = str(s.get("name") or "")
        ms = s.get("ms")
        ok = s.get("ok")
        ms_txt = f"{int(round(float(ms)))}ms" if isinstance(ms, (int, float)) else "N/A"
        ok_txt = "" if ok is None else (" ok" if ok else " fail")
        if name:
            out.append(f"{name}: {ms_txt}{ok_txt}")
    return out


def http_json(
    method: str,
    url: str,
    payload: Optional[Dict[str, Any]] = None,
    token: Optional[str] = None,
    timeout_sec: int = 45,
) -> Tuple[int, Dict[str, Any]]:
    data = None
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    req = request.Request(url, data=data, headers=headers, method=method)
    try:
        with request.urlopen(req, timeout=timeout_sec) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            if not body.strip():
                return resp.status, {}
            try:
                return resp.status, json.loads(body)
            except Exception:
                return resp.status, {"raw": body}
    except error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else ""
        try:
            parsed = json.loads(body) if body else {}
        except Exception:
            parsed = {"raw": body}
        return e.code, parsed


@dataclass
class BenchRef:
    sheet: str
    row: int
    row_no: str
    category: str
    lang_hint: str
    question: str
    expected_answer: str
    expected_source: str


@dataclass
class UniqueQuestion:
    question: str
    language: str
    refs: List[BenchRef] = field(default_factory=list)


class Runner:
    def __init__(
        self,
        xlsx_path: Path,
        out_dir: Path,
        api_base: str,
        poll_ms: int,
        timeout_s: int,
        dedupe: bool,
        max_questions: int,
        all_file_search: bool,
    ):
        self.xlsx_path = xlsx_path
        self.out_dir = out_dir
        self.api_base = api_base.rstrip("/")
        self.poll_ms = max(100, poll_ms)
        self.timeout_s = max(10, timeout_s)
        self.dedupe = dedupe
        self.max_questions = max_questions
        self.all_file_search = all_file_search
        self.token = ""

    def login(self) -> str:
        env_user = os.getenv("BENCH_EMPLOYEE_ID", "admin")
        env_pw = os.getenv("BENCH_PASSWORD", "")
        creds = []
        if env_pw:
            creds.append((env_user, env_pw))
        creds.extend([("admin", "password"), ("admin", "12345")])

        last = None
        for user, pw in creds:
            status, data = http_json(
                "POST",
                f"{self.api_base}/api/auth/login",
                {"employeeId": user, "password": pw},
            )
            token = ((data or {}).get("result") or {}).get("token")
            if status == 200 and token:
                self.token = str(token)
                return self.token
            last = (status, data)
        raise RuntimeError(f"login_failed {last}")

    def create_task(self, question: str) -> str:
        payload = {
            "type": "CHAT",
            "formData": {
                "prompt": question,
                "fieldSort": 1,
                "taskId": f"bench-detailed-{uuid.uuid4()}",
                "fileId": [0],
                "allFileSearch": self.all_file_search,
                "useMcp": False,
                "debug": True,
            },
        }
        status, data = http_json("POST", f"{self.api_base}/api/gen-task", payload, self.token)
        if status == 401:
            self.login()
            status, data = http_json("POST", f"{self.api_base}/api/gen-task", payload, self.token)

        task_id = ((data or {}).get("result") or {}).get("taskId")
        if not task_id:
            raise RuntimeError(f"task_create_error http={status} body={str(data)[:500]}")
        return str(task_id)

    def poll_task(self, task_id: str) -> Dict[str, Any]:
        started = time.time()
        cycles = 0
        while (time.time() - started) < self.timeout_s:
            cycles += 1
            url = (
                f"{self.api_base}/api/gen-task-output/list?"
                f"pageNum=1&pageSize=50&taskId={parse.quote(task_id)}"
            )
            status, data = http_json("GET", url, None, self.token)
            if status == 401:
                self.login()
                status, data = http_json("GET", url, None, self.token)

            rows = ((data or {}).get("result") or {}).get("rows") or []
            row = None
            if rows:
                row = next((r for r in rows if str(r.get("sort")) == "1"), rows[0])
            if row:
                st = str(row.get("status") or "")
                if st in ("FINISHED", "FAILED", "CANCEL"):
                    return {
                        "status": st,
                        "output_id": row.get("id"),
                        "content": str(row.get("content") or ""),
                        "elapsed_ms": int((time.time() - started) * 1000),
                        "poll_cycles": cycles,
                    }

            time.sleep(self.poll_ms / 1000.0)

        return {
            "status": "TIMEOUT",
            "output_id": None,
            "content": "",
            "elapsed_ms": int((time.time() - started) * 1000),
            "poll_cycles": cycles,
        }

    def fetch_kpi(self, output_id: Any) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        if not output_id:
            return None, None
        url = f"{self.api_base}/api/rag/kpi?outputId={parse.quote(str(output_id))}"
        status, data = http_json("GET", url, None, self.token)
        if status == 401:
            self.login()
            status, data = http_json("GET", url, None, self.token)
        if status != 200:
            return None, f"kpi_fetch_error http={status}"
        return (data or {}).get("data"), None

    def fetch_trace(self, task_id: str, output_id: Any) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        if not task_id or not output_id:
            return None, None
        url = (
            f"{self.api_base}/api/rag/trace?taskId={parse.quote(str(task_id))}"
            f"&outputId={parse.quote(str(output_id))}"
        )
        status, data = http_json("GET", url, None, self.token)
        if status == 401:
            self.login()
            status, data = http_json("GET", url, None, self.token)
        if status != 200:
            return None, f"trace_fetch_error http={status}"
        return (data or {}).get("data"), None

    def load_questions(self) -> Tuple[List[UniqueQuestion], Dict[str, List[BenchRef]], int]:
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        by_question: Dict[str, List[BenchRef]] = defaultdict(list)
        total_refs = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers: Dict[str, int] = {}
            for col in range(1, ws.max_column + 1):
                h = normalize_header(get_cell_text(ws.cell(row=1, column=col).value))
                if h and h not in headers:
                    headers[h] = col

            q_en_col = None
            q_ja_col = None
            q_ja_only_col = None
            for h, c in headers.items():
                if "question english" in h:
                    q_en_col = c
                elif "question japanese" in h:
                    q_ja_col = c
                elif "japanese question" in h:
                    q_ja_only_col = c

            source_col = headers.get("source")
            expected_col = headers.get("expected answer")
            category_col = headers.get("category")
            row_no_col = headers.get("no") or headers.get("id") or 1

            for r in range(2, ws.max_row + 1):
                row_no = get_cell_text(ws.cell(row=r, column=row_no_col).value)
                category = get_cell_text(ws.cell(row=r, column=category_col).value) if category_col else ""
                expected_answer = (
                    get_cell_text(ws.cell(row=r, column=expected_col).value) if expected_col else ""
                )
                expected_source = (
                    get_cell_text(ws.cell(row=r, column=source_col).value) if source_col else ""
                )

                candidates: List[Tuple[str, str]] = []
                if q_en_col:
                    q = normalize_whitespace(get_cell_text(ws.cell(row=r, column=q_en_col).value))
                    if q:
                        candidates.append((q, "en"))
                if q_ja_col:
                    q = normalize_whitespace(get_cell_text(ws.cell(row=r, column=q_ja_col).value))
                    if q:
                        candidates.append((q, "ja"))
                if q_ja_only_col:
                    q = normalize_whitespace(get_cell_text(ws.cell(row=r, column=q_ja_only_col).value))
                    if q:
                        candidates.append((q, "ja"))

                for question, lang in candidates:
                    ref = BenchRef(
                        sheet=sheet_name,
                        row=r,
                        row_no=row_no,
                        category=category,
                        lang_hint=lang,
                        question=question,
                        expected_answer=expected_answer,
                        expected_source=expected_source,
                    )
                    by_question[question].append(ref)
                    total_refs += 1

        unique_questions: List[UniqueQuestion] = []
        for q, refs in by_question.items():
            lang = "ja" if has_cjk(q) else "en"
            unique_questions.append(UniqueQuestion(question=q, language=lang, refs=refs))

        unique_questions.sort(key=lambda x: x.question)
        if self.max_questions > 0:
            unique_questions = unique_questions[: self.max_questions]
            allowed = {x.question for x in unique_questions}
            by_question = {k: v for k, v in by_question.items() if k in allowed}
            total_refs = sum(len(v) for v in by_question.values())

        return unique_questions, by_question, total_refs

    def run(self) -> Tuple[Path, Path]:
        self.out_dir.mkdir(parents=True, exist_ok=True)
        unique_questions, by_question, total_refs = self.load_questions()
        if not unique_questions:
            raise RuntimeError("no_questions_found")

        print(
            f"[load] workbook={self.xlsx_path} unique_questions={len(unique_questions)} "
            f"total_refs={total_refs} dedupe={self.dedupe}"
        )

        self.login()

        unique_results: Dict[str, Dict[str, Any]] = {}
        total_to_run = len(unique_questions) if self.dedupe else total_refs

        if self.dedupe:
            iterator = enumerate(unique_questions, start=1)
            for idx, uq in iterator:
                q = uq.question
                entry = self._execute_question(q)
                entry["language"] = uq.language
                entry["ref_count"] = len(uq.refs)
                entry["refs"] = [
                    {
                        "sheet": r.sheet,
                        "row": r.row,
                        "row_no": r.row_no,
                        "category": r.category,
                        "lang_hint": r.lang_hint,
                    }
                    for r in uq.refs
                ]
                unique_results[q] = entry
                marker = "IMPROPER" if entry.get("no_answer_reason") else "OK"
                print(
                    f"[{idx}/{total_to_run}] {marker} status={entry.get('status')} "
                    f"{entry.get('elapsed_ms')}ms refs={entry.get('ref_count')} :: {q[:110]}"
                )
        else:
            flat_index = 0
            for uq in unique_questions:
                for _ in uq.refs:
                    flat_index += 1
                    q = uq.question
                    entry = self._execute_question(q)
                    entry["language"] = uq.language
                    entry["ref_count"] = 1
                    entry["refs"] = []
                    unique_results[f"{q}##{flat_index}"] = entry
                    marker = "IMPROPER" if entry.get("no_answer_reason") else "OK"
                    print(
                        f"[{flat_index}/{total_to_run}] {marker} status={entry.get('status')} "
                        f"{entry.get('elapsed_ms')}ms :: {q[:110]}"
                    )

        flat_rows: List[Dict[str, Any]] = []
        for q, refs in by_question.items():
            result = unique_results.get(q)
            if not result:
                result = {
                    "question": q,
                    "status": "ERROR",
                    "answer": "",
                    "sources": [],
                    "elapsed_ms": 0,
                    "task_id": "",
                    "output_id": None,
                    "no_answer_reason": "missing_result",
                    "kpi": None,
                    "trace": None,
                    "poll_cycles": 0,
                    "inline_citations": False,
                }

            for ref in refs:
                eval_info = evaluate_answer(
                    ref.expected_answer,
                    result.get("answer") or "",
                    ref.expected_source,
                    result.get("sources") or [],
                )
                issue_codes: List[str] = []
                if result.get("no_answer_reason"):
                    issue_codes.append(str(result["no_answer_reason"]))
                if result.get("status") != "FINISHED":
                    issue_codes.append(f"status_{str(result.get('status') or '').lower()}")
                kpi = result.get("kpi") or {}
                rag_used = kpi.get("ragUsed")
                if rag_used is False:
                    issue_codes.append("rag_not_used")
                if not (result.get("sources") or []):
                    issue_codes.append("missing_source_citation")
                if eval_info.get("verdict") == "FAIL":
                    issue_codes.append("eval_low_semantic")
                if eval_info.get("source_match") == "NO":
                    issue_codes.append("expected_source_not_cited")

                deduped_issues = []
                seen_issues = set()
                for code in issue_codes:
                    if code and code not in seen_issues:
                        seen_issues.add(code)
                        deduped_issues.append(code)

                proper_output = len(deduped_issues) == 0

                flat_rows.append(
                    {
                        "sheet": ref.sheet,
                        "row": ref.row,
                        "row_no": ref.row_no,
                        "category": ref.category,
                        "lang_hint": ref.lang_hint,
                        "question": ref.question,
                        "expected_answer": ref.expected_answer,
                        "expected_source": ref.expected_source,
                        "status": result.get("status"),
                        "elapsed_ms": result.get("elapsed_ms"),
                        "poll_cycles": result.get("poll_cycles"),
                        "task_id": result.get("task_id"),
                        "output_id": result.get("output_id"),
                        "answer": result.get("answer"),
                        "sources": result.get("sources"),
                        "inline_citations": result.get("inline_citations"),
                        "no_answer_reason": result.get("no_answer_reason") or "",
                        "kpi": result.get("kpi"),
                        "trace": result.get("trace"),
                        "evaluation": eval_info,
                        "proper_output": proper_output,
                        "issues": deduped_issues,
                        "issue_explanations": [explain_reason(x) for x in deduped_issues],
                    }
                )

        summary = self._build_summary(unique_questions, total_refs, flat_rows)

        stamp = now_ts()
        json_path = self.out_dir / f"benchmark_qa_detailed_results_{stamp}.json"
        md_path = self.out_dir / f"benchmark_qa_detailed_report_{stamp}.md"
        latest_json = self.out_dir / "benchmark_qa_detailed_results_latest.json"
        latest_md = self.out_dir / "benchmark_qa_detailed_report_latest.md"

        payload = {
            "summary": summary,
            "unique_results": list(unique_results.values()),
            "rows": flat_rows,
            "metadata": {
                "generated_at_unix": int(time.time()),
                "generated_at_iso": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                "xlsx_path": str(self.xlsx_path),
                "api_base": self.api_base,
                "dedupe_unique_questions": self.dedupe,
            },
        }

        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        latest_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        md_text = self._build_markdown(summary, flat_rows, payload["metadata"])
        md_path.write_text(md_text, encoding="utf-8")
        latest_md.write_text(md_text, encoding="utf-8")

        print(f"[done] json={json_path}")
        print(f"[done] md={md_path}")
        print(
            f"[summary] proper={summary['proper_outputs']} improper={summary['improper_outputs']} "
            f"rows={summary['total_question_refs']} unique={summary['unique_questions']}"
        )
        return md_path, json_path

    def _execute_question(self, question: str) -> Dict[str, Any]:
        try:
            task_id = self.create_task(question)
        except Exception as e:
            return {
                "question": question,
                "status": "ERROR",
                "answer": "",
                "sources": [],
                "no_answer_reason": f"task_create_error: {e}",
                "elapsed_ms": 0,
                "poll_cycles": 0,
                "kpi": None,
                "trace": None,
                "task_id": "",
                "output_id": None,
                "inline_citations": False,
            }

        try:
            polled = self.poll_task(task_id)
        except Exception as e:
            return {
                "question": question,
                "status": "ERROR",
                "answer": "",
                "sources": [],
                "no_answer_reason": f"poll_error: {e}",
                "elapsed_ms": 0,
                "poll_cycles": 0,
                "kpi": None,
                "trace": None,
                "task_id": task_id,
                "output_id": None,
                "inline_citations": False,
            }

        answer = extract_answer(polled.get("content") or "")
        sources = extract_sources(answer)
        no_answer_reason = classify_no_answer(str(polled.get("status") or ""), answer)

        kpi, kpi_err = self.fetch_kpi(polled.get("output_id"))
        trace, trace_err = self.fetch_trace(task_id, polled.get("output_id"))
        if not no_answer_reason and kpi_err:
            no_answer_reason = kpi_err
        if not no_answer_reason and trace_err:
            no_answer_reason = trace_err

        return {
            "question": question,
            "status": polled.get("status"),
            "answer": answer,
            "sources": sources,
            "no_answer_reason": no_answer_reason,
            "elapsed_ms": polled.get("elapsed_ms"),
            "poll_cycles": polled.get("poll_cycles"),
            "kpi": kpi,
            "trace": trace,
            "task_id": task_id,
            "output_id": polled.get("output_id"),
            "inline_citations": has_inline_citations(answer),
        }

    def _build_summary(
        self,
        unique_questions: List[UniqueQuestion],
        total_refs: int,
        rows: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        proper = [r for r in rows if r.get("proper_output")]
        improper = [r for r in rows if not r.get("proper_output")]
        no_answer = [r for r in rows if r.get("no_answer_reason")]
        finished = [r for r in rows if r.get("status") == "FINISHED"]

        issue_counter = Counter()
        for r in rows:
            for issue in r.get("issues") or []:
                issue_counter[issue] += 1

        no_answer_counter = Counter()
        for r in no_answer:
            code = r.get("no_answer_reason") or "unknown"
            no_answer_counter[code] += 1

        sem_values = [
            int(r["evaluation"]["semantic_pct"])
            for r in rows
            if isinstance((r.get("evaluation") or {}).get("semantic_pct"), int)
        ]

        def avg_of(path: str) -> float:
            vals = []
            keys = path.split(".")
            for r in finished:
                cur = r
                ok = True
                for k in keys:
                    if not isinstance(cur, dict):
                        ok = False
                        break
                    cur = cur.get(k)
                if ok and isinstance(cur, (int, float)) and math.isfinite(float(cur)):
                    vals.append(float(cur))
            if not vals:
                return 0.0
            return float(sum(vals) / len(vals))

        rag_used_true = sum(
            1
            for r in finished
            if isinstance(r.get("kpi"), dict) and r["kpi"].get("ragUsed") is True
        )
        rag_used_false = sum(
            1
            for r in finished
            if isinstance(r.get("kpi"), dict) and r["kpi"].get("ragUsed") is False
        )

        by_category = Counter()
        for r in rows:
            by_category[(r.get("category") or "UNKNOWN")] += 1

        return {
            "unique_questions": len(unique_questions),
            "total_question_refs": total_refs,
            "finished": len(finished),
            "proper_outputs": len(proper),
            "improper_outputs": len(improper),
            "no_answer_count": len(no_answer),
            "issue_breakdown": dict(issue_counter.most_common()),
            "no_answer_breakdown": dict(no_answer_counter.most_common()),
            "avg_semantic_pct": round((sum(sem_values) / len(sem_values)), 2) if sem_values else None,
            "avg_elapsed_ms": round(avg_of("elapsed_ms"), 2),
            "avg_kpi_total_ms": round(avg_of("kpi.totalMs"), 2),
            "avg_kpi_rag_ms": round(avg_of("kpi.ragMs"), 2),
            "avg_kpi_retrieval_ms": round(avg_of("kpi.retrievalMs"), 2),
            "avg_kpi_llm_ms": round(avg_of("kpi.llmMs"), 2),
            "avg_kpi_translation_ms": round(avg_of("kpi.translationMs"), 2),
            "avg_input_tokens": round(avg_of("kpi.inputTokens"), 2),
            "avg_output_tokens": round(avg_of("kpi.outputTokens"), 2),
            "rag_used_true": rag_used_true,
            "rag_used_false": rag_used_false,
            "category_counts": dict(by_category.most_common()),
        }

    def _build_markdown(
        self,
        summary: Dict[str, Any],
        rows: List[Dict[str, Any]],
        metadata: Dict[str, Any],
    ) -> str:
        lines: List[str] = []
        lines.append("# Benchmark QA Detailed RAG Report")
        lines.append("")
        lines.append(f"- Generated At (UTC): **{metadata.get('generated_at_iso')}**")
        lines.append(f"- Workbook: **`{metadata.get('xlsx_path')}`**")
        lines.append(f"- API Base: **`{metadata.get('api_base')}`**")
        lines.append(f"- Dedupe Unique Questions: **{metadata.get('dedupe_unique_questions')}**")
        lines.append("")
        lines.append("## Summary")
        lines.append("")
        lines.append(f"- Unique questions executed: **{summary.get('unique_questions')}**")
        lines.append(f"- Total row references covered: **{summary.get('total_question_refs')}**")
        lines.append(f"- Finished responses: **{summary.get('finished')}**")
        lines.append(f"- Proper outputs: **{summary.get('proper_outputs')}**")
        lines.append(f"- Improper outputs: **{summary.get('improper_outputs')}**")
        lines.append(f"- No-answer count: **{summary.get('no_answer_count')}**")
        lines.append(f"- Avg semantic score (%): **{summary.get('avg_semantic_pct')}**")
        lines.append("")
        lines.append("### KPI Averages")
        lines.append("")
        lines.append(f"- Avg end-to-end elapsed: **{fmt_ms(summary.get('avg_elapsed_ms'))}**")
        lines.append(f"- Avg KPI totalMs: **{fmt_ms(summary.get('avg_kpi_total_ms'))}**")
        lines.append(f"- Avg KPI ragMs: **{fmt_ms(summary.get('avg_kpi_rag_ms'))}**")
        lines.append(f"- Avg KPI retrievalMs: **{fmt_ms(summary.get('avg_kpi_retrieval_ms'))}**")
        lines.append(f"- Avg KPI llmMs: **{fmt_ms(summary.get('avg_kpi_llm_ms'))}**")
        lines.append(f"- Avg KPI translationMs: **{fmt_ms(summary.get('avg_kpi_translation_ms'))}**")
        lines.append(f"- Avg input tokens: **{summary.get('avg_input_tokens')}**")
        lines.append(f"- Avg output tokens: **{summary.get('avg_output_tokens')}**")
        lines.append(f"- ragUsed=true rows: **{summary.get('rag_used_true')}**")
        lines.append(f"- ragUsed=false rows: **{summary.get('rag_used_false')}**")
        lines.append("")

        lines.append("### Improper Output Breakdown")
        lines.append("")
        for code, count in (summary.get("issue_breakdown") or {}).items():
            lines.append(f"- `{code}`: **{count}** - {explain_reason(code)}")
        if not (summary.get("issue_breakdown") or {}):
            lines.append("- None")
        lines.append("")

        lines.append("### No-Answer Breakdown")
        lines.append("")
        for code, count in (summary.get("no_answer_breakdown") or {}).items():
            lines.append(f"- `{code}`: **{count}** - {explain_reason(code)}")
        if not (summary.get("no_answer_breakdown") or {}):
            lines.append("- None")
        lines.append("")

        lines.append("## Detailed Results (By Sheet Row Reference)")
        lines.append("")

        for i, row in enumerate(rows, start=1):
            q = row.get("question") or ""
            lines.append(f"### {i}. [{row.get('sheet')} r{row.get('row')}] {q}")
            lines.append(
                f"- Row No: `{row.get('row_no')}` | Category: `{row.get('category')}` | Lang: `{row.get('lang_hint')}`"
            )
            lines.append(
                f"- Task: `{row.get('task_id')}` | Output: `{row.get('output_id')}` | Status: `{row.get('status')}`"
            )
            lines.append(
                f"- Proper Output: **{'YES' if row.get('proper_output') else 'NO'}** | "
                f"Elapsed: `{fmt_ms(row.get('elapsed_ms'))}` | Poll cycles: `{row.get('poll_cycles')}`"
            )
            if row.get("no_answer_reason"):
                reason = row.get("no_answer_reason")
                lines.append(f"- No-answer reason: `{reason}` - {explain_reason(reason)}")

            ev = row.get("evaluation") or {}
            lines.append(
                "- Evaluation: "
                f"verdict=`{ev.get('verdict')}`, "
                f"semantic=`{ev.get('semantic_pct')}`, "
                f"recall=`{ev.get('recall_pct')}`, "
                f"jaccard=`{ev.get('jaccard_pct')}`, "
                f"source_match=`{ev.get('source_match')}`"
            )

            issues = row.get("issues") or []
            if issues:
                lines.append("- Improper-output issues:")
                for code in issues:
                    lines.append(f"  - `{code}`: {explain_reason(code)}")

            lines.append(f"- Expected Source: `{row.get('expected_source') or ''}`")
            lines.append("- Expected Answer:")
            lines.append("```text")
            lines.append(str(row.get("expected_answer") or ""))
            lines.append("```")

            lines.append("- Actual Answer:")
            lines.append("```text")
            lines.append(str(row.get("answer") or ""))
            lines.append("```")

            srcs = row.get("sources") or []
            lines.append(f"- Extracted Sources: `{', '.join(srcs) if srcs else ''}`")
            lines.append(f"- Inline citations present: `{row.get('inline_citations')}`")

            kpi = row.get("kpi") or {}
            lines.append("- KPI Metrics:")
            lines.append(f"  - totalMs: `{fmt_ms(kpi.get('totalMs'))}`")
            lines.append(f"  - ragMs: `{fmt_ms(kpi.get('ragMs'))}`")
            lines.append(f"  - retrievalMs: `{fmt_ms(kpi.get('retrievalMs'))}`")
            lines.append(f"  - llmMs: `{fmt_ms(kpi.get('llmMs'))}`")
            lines.append(f"  - translationMs: `{fmt_ms(kpi.get('translationMs'))}`")
            lines.append(f"  - queryTranslationMs: `{fmt_ms(kpi.get('queryTranslationMs'))}`")
            lines.append(f"  - titleMs: `{fmt_ms(kpi.get('titleMs'))}`")
            lines.append(f"  - inputTokens: `{kpi.get('inputTokens')}`")
            lines.append(f"  - outputTokens: `{kpi.get('outputTokens')}`")
            lines.append(f"  - ragUsed: `{kpi.get('ragUsed')}`")
            lines.append(f"  - userLanguage: `{kpi.get('userLanguage')}`")

            trace = row.get("trace") or {}
            lines.append("- Trace Metrics:")
            lines.append(f"  - totalMs: `{fmt_ms(trace.get('totalMs'))}`")
            lines.append(f"  - ttftMs: `{fmt_ms(trace.get('ttftMs'))}`")
            stage_summaries = summarize_stage_list(trace)
            if stage_summaries:
                lines.append("  - stages:")
                for st in stage_summaries:
                    lines.append(f"    - {st}")
            else:
                lines.append("  - stages: `N/A`")

            lines.append("")

        return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run Benchmark QA workbook against local RAG API and build detailed markdown+json reports.",
    )
    parser.add_argument(
        "--xlsx",
        default="/home/qabot/hrbot/Benchmark QA.xlsx",
        help="Path to Benchmark QA workbook.",
    )
    parser.add_argument(
        "--out-dir",
        default="/home/qabot/hrbot",
        help="Directory for output markdown/json files.",
    )
    parser.add_argument(
        "--api-base",
        default="http://127.0.0.1:8080",
        help="Base URL of API server.",
    )
    parser.add_argument(
        "--poll-ms",
        type=int,
        default=700,
        help="Polling interval in ms for task output.",
    )
    parser.add_argument(
        "--timeout-s",
        type=int,
        default=180,
        help="Timeout per question in seconds.",
    )
    parser.add_argument(
        "--max-questions",
        type=int,
        default=0,
        help="Optional cap on unique questions (0 = all).",
    )
    parser.add_argument(
        "--no-dedupe",
        action="store_true",
        help="Disable unique-question dedupe; executes every row reference separately.",
    )
    parser.add_argument(
        "--all-file-search",
        type=int,
        default=1,
        help="Whether to set allFileSearch=true (1) or false (0).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    runner = Runner(
        xlsx_path=Path(args.xlsx).resolve(),
        out_dir=Path(args.out_dir).resolve(),
        api_base=str(args.api_base),
        poll_ms=args.poll_ms,
        timeout_s=args.timeout_s,
        dedupe=not args.no_dedupe,
        max_questions=max(0, int(args.max_questions)),
        all_file_search=bool(int(args.all_file_search)),
    )

    md_path, json_path = runner.run()
    print(f"[final] markdown={md_path}")
    print(f"[final] json={json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
